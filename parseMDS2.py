import argparse
import xml.etree.ElementTree as eTree
import re
import os
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description='Parse MDS2 HN-2019 XLSX into XML')
parser.add_argument('infile', metavar='*.xlsx', help='MDS2 form in .xls or .xlsx format')

args = parser.parse_args()


def indent_xml(elem, level=0):
    i = "\n" + level*"  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent_xml(elem, level+1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


def get_all_rows(data_file):
    all_rows = []
    wb = load_workbook(data_file, data_only=True)

    for sheet in wb.sheetnames:
        try:
            if re.search('mds2', sheet, re.IGNORECASE):
                ws = wb[sheet]
                all_rows = list(ws.rows)
            elif re.search('Sheet1', sheet, re.IGNORECASE):
                ws = wb[sheet]
                all_rows = list(ws.rows)
            else:
                print('Unable to determine sheet containing MDS2')

        except (AttributeError, TypeError):
            continue

    return all_rows


def get_notes(questionid):
    note = ''
    all_rows = get_all_rows(args.infile)

    for row in all_rows:
        for any_cell in row:
            try:
                if any_cell.value == questionid:
                    note = str(row[1].value)

            except (AttributeError, TypeError):
                continue

    return note


def generate_xml():
    root = eTree.Element('mds2')

    sections = eTree.Element('sections')
    section_doc = eTree.SubElement(sections, 'section')
    section_doc.set('name', 'DOCUMENT')
    section_mpii = eTree.SubElement(sections, 'section')
    section_mpii.set('name', 'MANAGEMENT OF PERSONALLY IDENTIFIABLE INFORMATION')
    section_alof = eTree.SubElement(sections, 'section')
    section_alof.set('name', 'AUTOMATIC LOGOFF')
    section_audt = eTree.SubElement(sections, 'section')
    section_audt.set('name', 'AUDIT CONTROLS')
    section_auth = eTree.SubElement(sections, 'section')
    section_auth.set('name', 'AUTHORIZATION')
    section_csup = eTree.SubElement(sections, 'section')
    section_csup.set('name', 'CYBER SECURITY PRODUCT UPGRADES')
    section_didt = eTree.SubElement(sections, 'section')
    section_didt.set('name', 'HEALTH DATA DE-IDENTIFICATION')
    section_dtbk = eTree.SubElement(sections, 'section')
    section_dtbk.set('name', 'DATA BACKUP AND DISASTER RECOVERY')
    section_emrg = eTree.SubElement(sections, 'section')
    section_emrg.set('name', 'EMERGENCY ACCESS')
    section_igau = eTree.SubElement(sections, 'section')
    section_igau.set('name', 'HEALTH DATA INTEGRITY AND AUTHENTICITY')
    section_mldp = eTree.SubElement(sections, 'section')
    section_mldp.set('name', 'MALWARE DETECTION / PROTECTION')
    section_naut = eTree.SubElement(sections, 'section')
    section_naut.set('name', 'NODE AUTHENTICATION')
    section_conn = eTree.SubElement(sections, 'section')
    section_conn.set('name', 'CONNECTIVITY CAPABILITIES')
    section_paut = eTree.SubElement(sections, 'section')
    section_paut.set('name', 'PERSON AUTHENTICATION')
    section_plok = eTree.SubElement(sections, 'section')
    section_plok.set('name', 'PHYSICAL LOCKS')
    section_rdmp = eTree.SubElement(sections, 'section')
    section_rdmp.set('name', 'ROADMAP FOR THIRD PARTY COMPONENTS IN DEVICE LIFE CYCLE')
    section_sbom = eTree.SubElement(sections, 'section')
    section_sbom.set('name', 'SOFTWARE BILL OF MATERIALS')
    section_sahd = eTree.SubElement(sections, 'section')
    section_sahd.set('name', 'SYSTEM AND APPLICATION HARDENING')
    section_sgud = eTree.SubElement(sections, 'section')
    section_sgud.set('name', 'SECURITY GUIDANCE')
    section_stcf = eTree.SubElement(sections, 'section')
    section_stcf.set('name', 'HEALTH DATA STORAGE CONFIDENTIALITY')
    section_txcf = eTree.SubElement(sections, 'section')
    section_txcf.set('name', 'TRANSMISSION CONFIDENTIALITY')
    section_txig = eTree.SubElement(sections, 'section')
    section_txig.set('name', 'TRANSMISSION INTEGRITY')
    section_rmot = eTree.SubElement(sections, 'section')
    section_rmot.set('name', 'REMOTE SERVICE')

    root.append(sections)

    questions_doc = eTree.SubElement(section_doc, 'questions')
    questions_mpii = eTree.SubElement(section_mpii, 'questions')
    questions_alof = eTree.SubElement(section_alof, 'questions')
    questions_audt = eTree.SubElement(section_audt, 'questions')
    questions_auth = eTree.SubElement(section_auth, 'questions')
    questions_csup = eTree.SubElement(section_csup, 'questions')
    questions_didt = eTree.SubElement(section_didt, 'questions')
    questions_dtbk = eTree.SubElement(section_dtbk, 'questions')
    questions_emrg = eTree.SubElement(section_emrg, 'questions')
    questions_igau = eTree.SubElement(section_igau, 'questions')
    questions_mldp = eTree.SubElement(section_mldp, 'questions')
    questions_naut = eTree.SubElement(section_naut, 'questions')
    questions_conn = eTree.SubElement(section_conn, 'questions')
    questions_paut = eTree.SubElement(section_paut, 'questions')
    questions_plok = eTree.SubElement(section_plok, 'questions')
    questions_rdmp = eTree.SubElement(section_rdmp, 'questions')
    questions_sbom = eTree.SubElement(section_sbom, 'questions')
    questions_sahd = eTree.SubElement(section_sahd, 'questions')
    questions_sgud = eTree.SubElement(section_sgud, 'questions')
    questions_stcf = eTree.SubElement(section_stcf, 'questions')
    questions_txcf = eTree.SubElement(section_txcf, 'questions')
    questions_txig = eTree.SubElement(section_txig, 'questions')
    questions_rmot = eTree.SubElement(section_rmot, 'questions')

    summarize_data(questions_doc, 'DOC-')
    summarize_data(questions_mpii, 'MPII-')
    summarize_data(questions_alof, 'ALOF-')
    summarize_data(questions_audt, 'AUDT-')
    summarize_data(questions_auth, 'AUTH-')
    summarize_data(questions_csup, 'CSUP-')
    summarize_data(questions_didt, 'DIDT-')
    summarize_data(questions_dtbk, 'DTBK-')
    summarize_data(questions_emrg, 'EMRG-')
    summarize_data(questions_igau, 'IGAU-')
    summarize_data(questions_mldp, 'MLDP-')
    summarize_data(questions_naut, 'NAUT-')
    summarize_data(questions_conn, 'CONN-')
    summarize_data(questions_paut, 'PAUT-')
    summarize_data(questions_plok, 'PLOK-')
    summarize_data(questions_rdmp, 'RDMP-')
    summarize_data(questions_sbom, 'SBOM-')
    summarize_data(questions_sahd, 'SAHD-')
    summarize_data(questions_sgud, 'SGUD-')
    summarize_data(questions_stcf, 'STCF-')
    summarize_data(questions_txcf, 'TXCF-')
    summarize_data(questions_txig, 'TXIG-')
    summarize_data(questions_rmot, 'RMOT-')

    indent_xml(root)
    tree = eTree.ElementTree(root)

    file_name = os.path.basename(args.infile)
    file_path, new_file = os.path.split(args.infile)
    new_file = os.path.splitext(new_file)[0] + '.xml'

    with open(file_path + '/' + new_file, 'wb') as files:
        tree.write(files, xml_declaration=True, encoding='utf-8', method='xml')


def summarize_data(section, abbreviation):
    all_rows = get_all_rows(args.infile)

    for row in all_rows:
        for any_cell in row:

            try:
                if any_cell.value.startswith(abbreviation):
                    doc_qid = eTree.SubElement(section, 'questionID')
                    doc_qid.set('id', str(row[0].value))
                    eTree.SubElement(doc_qid, 'question').text = str(row[1].value)
                    eTree.SubElement(doc_qid, 'response').text = str(row[2].value)

                    try:
                        if str(row[0].value) in str(row[3].value):
                            eTree.SubElement(doc_qid, 'note').text = get_notes(str(row[3].value))
                        else:
                            eTree.SubElement(doc_qid, 'note').text = str(row[3].value)

                    except (AttributeError, TypeError):
                        continue

            except (AttributeError, TypeError):
                continue


generate_xml()